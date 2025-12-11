import openpyxl
import pandas as pd
import os
import datetime

"""
读取csv文件写入excel表格 
"""

now = datetime.datetime.now()
end = datetime.datetime(now.year, now.month, 1) - datetime.timedelta(days=1)
L_upDate = str(end.year) + "." + str(end.month - 1)
if end.month >= 10:
	upDate = str(end.year) + "." + str(end.month)
else:
	upDate = str(end.year) + ".0" + str(end.month)


def csv2excel():
	# 创建工作簿对象
	work_book = openpyxl.Workbook()
	# 创建sheet
	work_sheet = work_book.create_sheet(index=0, title="Summary")
	# 读取csv文件
	try:
		content = pd.read_csv('QC_summary.csv').values.tolist()
	except pd.errors.EmptyDataError:
		content = []
	# 写入第一行，标题
	work_sheet["A1"] = "Summary"
	# 从第二行开始写入从csv读取的内容
	row = 2
	column = 'A'
	for line in content:
		for text in line:
			work_sheet[column + str(row)] = text
		row += 1
	os.remove('QC_summary.csv')

	'''
	创建sheet
	work_sheet = work_book.create_sheet(index=1, title="Symbol")
	try:
		content = pd.read_csv('QC_symbol.csv').values.tolist()
	except pd.errors.EmptyDataError:
		content = []
	写入第一行，标题
	work_sheet["A1"] = "Detail Symbol"
	从第二行开始写入从csv读取的内容
	row = 2
	column = "A"
	for line in content:
		for text in line:
			work_sheet[column + str(row)] = text
		row += 1
	os.remove('QC_symbol.csv')
	'''

	# 创建sheet
	work_sheet = work_book.create_sheet(index=2, title="Data")
	sub_content = []
	try:
		content = pd.read_csv('QC_data.csv').values.tolist()
	except pd.errors.EmptyDataError:
		content = []

	end_index = 0
	start_index = 0
	count = 1
	num = 0
	symbol = content[0][1]

	# 超过20条数据隐藏
	while num < len(content):
		if content[num][1] == symbol:  # symbol相同
			if count < 20:  # 数量没有达到20
				count += 1
			elif count == 20:  # 数量达到20
				# 如果是最后一个symbol, 结束循环
				if content[num][1] == content[-1][1]:
					end_index = len(content) - 1
					sub_content.extend(content[end_index - 20: end_index])
					sub_content.append([content[num][0], content[num][1] + "  20+ Records"])
					break
				else:
					# 找出下一symbol的下标
					for i in range(num + 1, len(content)):
						if content[i][1] != symbol:
							start_index = i
							end_index = i - 1
							symbol = content[i][1]
							num = i - 1
							break
					sub_content.extend(content[end_index - 19: end_index + 1])
					sub_content.append([content[num][0], content[num][1] + "  20+ Records"])
					count = 1
		else:  # 数量少于20 and symbol更改
			sub_content.extend(content[start_index: num])
			start_index = num
			symbol = content[num][1]
			count = 1
		num += 1

	# 写入第一行，标题
	columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
	title_list = ['Nation', 'Symbol', 'Date', 'Open', 'High', 'Low', 'Close']
	for i, column in enumerate(columns):
		if i < 7:
			work_sheet[column + "2"] = title_list[i]
		else:
			work_sheet[column + "2"] = title_list[i-4]
	work_sheet["D1"] = upDate
	work_sheet["H1"] = L_upDate

	# 从第三行开始写入从csv读取的内容
	row = 3
	for line in sub_content:
		column = 0
		for text in line:
			work_sheet[columns[column] + str(row)] = text
			column += 1
		row += 1
	os.remove('QC_data.csv')
	# 保存工作表
	work_book.remove(work_book.get_sheet_by_name("Sheet"))
	work_book.save('QC_report_' + upDate + '.xlsx')
